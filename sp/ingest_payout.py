#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Ingest any Zomato settlement workbook -> normalized monthly payout dict.

Robust to the 3 column layouts seen across report versions: columns are mapped
by NORMALIZED HEADER NAME, not position. Handles the fee-model differences
(ld / fulfilment / none), variable row widths, #REF! rows, and duplicate files.
Also pulls real ad spend from the 'Addition Deductions Details' sheet.

Returns a dict per restaurant-month (see schema in module docstring).
"""
import glob
import os
import re
from datetime import datetime

import openpyxl


def norm(name):
    s = name.lower().replace("[", " ").replace("]", " ").replace("(", " ").replace(")", " ")
    s = s.replace(".", " ")  # "Res. name" -> "res name"
    return re.sub(r"\s+", " ", s).strip()


def find_col(cols, *fragments):
    for frag in fragments:
        for cname, cnum in cols.items():
            if frag in cname:
                return cnum
    return None


def find_fee_col(cols):
    """Commission/service-fee column: the fee itself, NOT the '%' column and
    NOT the '& payment mechanism fees' aggregate. A trailing '%' marks the
    percentage column; a '%' inside a formula (e.g. '[(12)% * (B)]') does not."""
    best = None
    for cname, cnum in cols.items():
        if cname.endswith("%") or "payment" in cname or "&" in cname:
            continue
        if "base service fee" in cname or "service fee" in cname:
            if best is None or cnum < best:
                best = cnum
    return best


def find_fulfil_col(cols):
    for cname, cnum in cols.items():
        if "fulfilment fee" in cname and "per km" not in cname and "calculation" not in cname and "distance" not in cname:
            return cnum
    return None


def fval(row, col):
    if col is None:
        return 0.0
    try:
        return float(row[col - 1])
    except (TypeError, ValueError, IndexError):
        return 0.0


def sval(row, col):
    if col is None:
        return ""
    try:
        return str(row[col - 1] or "").strip()
    except IndexError:
        return ""


def _month_of(daily):
    if not daily:
        return ""
    d = min(daily.keys())
    try:
        return datetime.strptime(d, "%Y-%m-%d").strftime("%Y-%m")
    except Exception:
        return d[:7]


def _max_days(daily):
    import calendar
    d = min(daily.keys())
    try:
        dt = datetime.strptime(d, "%Y-%m-%d")
        return calendar.monthrange(dt.year, dt.month)[1]
    except Exception:
        return 30


def scan_payout(path):
    """Return a normalized monthly payout dict for one settlement workbook."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Order Level"]
    hdr_row = None
    for r in range(1, 12):
        if any(c.value and "Subtotal" in str(c.value) for c in ws[r]):
            hdr_row = r
            break
    if hdr_row is None:
        wb.close()
        return None
    cols = {}
    for c in ws[hdr_row]:
        if c.value:
            cols[norm(str(c.value))] = c.column
    C = dict(
        oid=find_col(cols, "order id"), date=find_col(cols, "order date"),
        status=find_col(cols, "order status"), res=find_col(cols, "res name"),
        resid=find_col(cols, "res id"), subtotal=find_col(cols, "subtotal"),
        packaging=find_col(cols, "packaging"),
        promo=find_col(cols, "restaurant discount promo"),
        bogo=find_col(cols, "restaurant discount bogo"),
        deldisc=find_col(cols, "delivery charge discount"),
        gst=find_col(cols, "total gst collected"), nov=find_col(cols, "net order value"),
        comm=find_fee_col(cols), dist=find_col(cols, "actual order distance"),
        ld=find_col(cols, "long distance enablement fee"), fulfil=find_fulfil_col(cols),
        pm=find_col(cols, "payment mechanism fee"), taxfees=find_col(cols, "taxes on service"),
        tds=find_col(cols, "tds 194o"), payout=find_col(cols, "order level payout"),
    )
    fee_model = "ld" if C["ld"] else ("fulfilment" if C["fulfil"] else "none")
    dist_fee_col = C["ld"] or C["fulfil"]
    if C["res"] is None:
        wb.close()
        return None

    t = dict(orders=0, cancelled=0, subtotal=0.0, packaging=0.0, gst=0.0, nov=0.0,
             commission=0.0, dist_fee=0.0, dist_fee_orders=0, pm=0.0, tax_fees=0.0,
             tds=0.0, payout=0.0, promo_disc=0.0, bogo_disc=0.0, deldisc=0.0,
             dist_km=0.0, promo_orders=0)
    hourly = {}
    daily = {}
    res_name = res_id = None
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        oid = sval(row, C["oid"])
        if not oid or oid.startswith("#"):
            continue
        st = sval(row, C["status"]).lower()
        if st in ("cancelled", "rejected"):
            t["cancelled"] += 1
            continue
        if st != "delivered":
            continue
        if res_name is None:
            res_name = sval(row, C["res"])
            res_id = sval(row, C["resid"])
        t["orders"] += 1
        t["subtotal"] += fval(row, C["subtotal"])
        t["packaging"] += fval(row, C["packaging"])
        t["gst"] += fval(row, C["gst"])
        t["nov"] += fval(row, C["nov"])
        t["commission"] += fval(row, C["comm"])
        dfv = fval(row, dist_fee_col)
        t["dist_fee"] += dfv
        if dfv > 0:
            t["dist_fee_orders"] += 1
        t["pm"] += fval(row, C["pm"])
        t["tax_fees"] += fval(row, C["taxfees"])
        t["tds"] += fval(row, C["tds"])
        t["payout"] += fval(row, C["payout"])
        t["promo_disc"] += fval(row, C["promo"])
        t["bogo_disc"] += fval(row, C["bogo"])
        t["deldisc"] += fval(row, C["deldisc"])
        t["dist_km"] += fval(row, C["dist"])
        if fval(row, C["promo"]) > 0 or fval(row, C["bogo"]) > 0:
            t["promo_orders"] += 1
        dv = row[C["date"] - 1] if C["date"] else None
        try:
            if isinstance(dv, datetime):
                dt = dv
            else:
                s = str(dv or "").strip()
                if len(s) < 19:
                    s = s + " 00:00:00"
                dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        hourly[dt.strftime("%H")] = hourly.get(dt.strftime("%H"), 0) + 1
        daily[dt.strftime("%Y-%m-%d")] = daily.get(dt.strftime("%Y-%m-%d"), 0) + 1
    wb.close()
    if t["orders"] == 0:
        return None

    # real ad spend from Addition Deductions Details
    ad_spend = 0.0
    ad_lines = 0
    try:
        wb2 = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if "Addition Deductions Details" in wb2.sheetnames:
            ws2 = wb2["Addition Deductions Details"]
            for row in ws2.iter_rows(values_only=True):
                if row and len(row) > 6 and row[1] and str(row[1]).strip().upper() == "ADS":
                    try:
                        ad_spend += float(row[6] or 0)
                    except (TypeError, ValueError, IndexError):
                        pass
                    ad_lines += 1
        wb2.close()
    except Exception:
        pass

    n = t["orders"]
    fees = t["commission"] + t["dist_fee"] + t["pm"] + t["tax_fees"] + t["tds"]
    dinner = sum(v for k, v in hourly.items() if k in ("19", "20", "21"))
    lunch = sum(v for k, v in hourly.items() if k in ("12", "13", "14"))
    return {
        "file": os.path.basename(path),
        "restaurant": res_name, "res_id": res_id, "month": _month_of(daily),
        "orders": n, "cancelled": t["cancelled"],
        "subtotal": round(t["subtotal"], 2), "packaging": round(t["packaging"], 2),
        "gst": round(t["gst"], 2), "nov": round(t["nov"], 2),
        "commission": round(t["commission"], 2), "dist_fee": round(t["dist_fee"], 2),
        "dist_fee_orders": t["dist_fee_orders"], "dist_fee_model": fee_model,
        "payment_mech": round(t["pm"], 2), "tax_on_fees": round(t["tax_fees"], 2),
        "tds": round(t["tds"], 2), "payout": round(t["payout"], 2),
        "promo_disc": round(t["promo_disc"], 2), "bogo_disc": round(t["bogo_disc"], 2),
        "delivery_disc": round(t["deldisc"], 2),
        "avg_distance_km": round(t["dist_km"] / n, 2) if n else 0,
        "aov": round(t["subtotal"] / n, 2) if n else 0,
        "take_rate_pct": round(100 * fees / t["subtotal"], 1) if t["subtotal"] else 0,
        "dist_fee_exposure_pct": round(100 * t["dist_fee_orders"] / n, 1) if n else 0,
        "promo_share_pct": round(100 * t["promo_orders"] / n, 1) if n else 0,
        "discount_rate_pct": round(100 * (t["promo_disc"] + t["bogo_disc"]) / t["subtotal"], 1) if t["subtotal"] else 0,
        "payout_pct_of_nov": round(100 * t["payout"] / t["nov"], 1) if t["nov"] else 0,
        "ad_spend_deductions": round(ad_spend, 2), "ad_lines": ad_lines,
        "dinner_pct": round(100 * dinner / n, 1) if n else 0,
        "lunch_pct": round(100 * lunch / n, 1) if n else 0,
        "zero_order_days": (len(daily) and (_max_days(daily) - len(daily))),
    }


def scan_folder(folder):
    """Scan a folder of payout xlsx -> list of monthly payout dicts (deduped)."""
    out = []
    seen = set()
    for f in sorted(glob.glob(os.path.join(folder, "*.xlsx"))):
        try:
            r = scan_payout(f)
        except Exception:
            r = None
        if r:
            key = (r["res_id"], r["month"])
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
    return out