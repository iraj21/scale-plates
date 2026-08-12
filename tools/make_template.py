#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Create the Scale Plates onboarding Excel template.

Sheets:
  HowTo        — 3-step instructions (the client-facing page)
  PayoutDump   — paste the Zomato 'Order Level' tab here (all rows)
  FunnelDump   — paste the funnel CSV(s) here (all rows, one after another)
  Insights     — filled by build_onboarding.py with KPIs + actionables

The dump sheets ship with a small SAMPLE so the flow can be tested instantly;
the HowTo says to replace with the client's own data.
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "1F4E79"
NAVY_D = "143352"
ORANGE = "E67E22"
GREEN = "2E7D32"
RED = "C0392B"
LIGHT = "F4F6F8"
GREY = "6B7280"
WHITE = "FFFFFF"
GOLD = "FFC107"

thin = Side(style="thin", color="D5DDE5")


def style_sheet(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False


def title_band(ws, title, sub, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, size=20, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY_D)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 34
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=10)
    c2 = ws.cell(row=row + 1, column=1, value=sub)
    c2.font = Font(size=10, color=GREY)
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row + 1].height = 30


def sample_order_rows():
    """A few rows of the Zomato Order Level tab (older layout) to demo the paste."""
    hdr = ["S.no.", "Order ID", "Order date", "Week no.", "Res. name", "Res. ID",
           "Discount construct", "Mode of payment", "Order status (Delivered/ Cancelled/ Rejected)",
           "Cancellation policy", "Cancellation/ Rejection reason", "Cancelled/ Rejected State",
           "Order type", "Delivery state code", "Subtotal (items total)", "Packaging charge",
           "Delivery charge for restaurants on self logistics", "Restaurant discount (Promo)",
           "Restaurant discount (BOGO, Freebies, Gold, Brand pack & other)", "Brand pack subscription fee",
           "Delivery charge discount/ Relisting discount", "Total GST collected from customers",
           "Net order value", "Commissionable value of Subtotal, excluding restaurant discount",
           "Commissionable value of Packaging charge", "Commissionable value of Total GST collected from customers",
           "Total commissionable value", "Base service fee %", "Base service fee",
           "Actual order distance (km)", "Long distance enablement fee", "Discount on long distance enablement fee",
           "Discount on service fee due to 30% capping", "Payment mechanism fee",
           "Service fee & payment mechanism fee", "Taxes on service fee & payment mechanism fee",
           "Applicable amount for TCS", "Applicable amount for 9(5)", "Tax collected at source",
           "TCS IGST amount", "TDS 194O amount", "GST paid by Zomato on behalf of restaurant",
           "GST to be paid by Restaurant partner to Govt.", "Government charges",
           "Customer compensation/ recoupment", "Delivery charges recovery",
           "Amount received in cash (on self delivery orders)", "Credit note/ (Debit note) adjustment",
           "Promo recovery adjustment", "Extra inventory ads (order level deduction)",
           "Brand loyalty points redemption", "Express order fee", "Other order-level deductions",
           "Net Deductions", "Net Additions", "Order level Payout", "Settlement status",
           "Settlement date", "Bank UTR", "Unsettled amount", "Customer ID"]
    r1 = [1, 7228691197, "2026-07-01 12:34:56", 27, "Kubaba", 21206855, "x", "Online",
          "Delivered", "", "", "", "Delivery", "32", 560.0, 0, 0, 0, 0, 0, 0, 28.47,
          588.47, 560.0, 0, 28.47, 588.47, 22, 123.2, 3.5, 0, 0, 0, 10.83, 134.03,
          24.13, 0, 0, 0, 0, 0.59, 0, 0, 24.72, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0,
          453.13, "Settled", "2026-07-03", "UTR123", 0, "C1001"]
    r2 = [2, 7226382301, "2026-07-01 13:10:00", 27, "Kubaba", 21206855, "x", "Online",
          "Delivered", "", "", "", "Delivery", "32", 170.0, 0, 0, 0, 0, 0, 0, 8.64,
          178.64, 170.0, 0, 8.64, 178.64, 22, 37.4, 2.1, 0, 0, 0, 3.29, 40.69, 7.32,
          0, 0, 0, 0, 0.18, 0, 0, 7.5, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 138.17,
          "Settled", "2026-07-03", "UTR124", 0, "C1002"]
    return hdr, [r1, r2]


def sample_funnel_rows():
    """A few rows of the daily funnel CSV to demo the paste."""
    hdr = ["Restaurant ID", "Restaurant name", "Subzone", "City", "Overview", "Metric",
           "01 Jul, 2026", "02 Jul, 2026", "03 Jul, 2026"]
    rows = [
        [21206855, "Kubaba", "Edappally", "Kochi", "Sales", "Delivered orders", 128, 81, 49],
        [21206855, "Kubaba", "Edappally", "Kochi", "Sales", "Sales (Rs)", 57730.49, 45763, 24234],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer funnel", "Impressions", 4810, 4275, 5681],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer funnel", "Menu opens", 1109, 1110, 1071],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer funnel", "Cart builds", 507, 512, 416],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer funnel", "Placed Orders", 128, 81, 49],
        [21206855, "Kubaba", "Edappally", "Kochi", "Ads", "Ads spend (Rs)", 4670, 3933, 3881],
        [21206855, "Kubaba", "Edappally", "Kochi", "Ads", "Sales from ads (Rs)", 20710.04, 17690, 14340],
        [21206855, "Kubaba", "Edappally", "Kochi", "Ads", "Ads orders", 53, 35, 33],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer experience", "Average rating", 4.248, 4.229, 4.246],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer segmentation", "Repeat user orders", 55, 25, 11],
        [21206855, "Kubaba", "Edappally", "Kochi", "Customer segmentation", "Dinner orders", 71, 29, 23],
    ]
    return hdr, rows


def build(path):
    wb = Workbook()

    # ---- HowTo ----
    ws = wb.active
    ws.title = "HowTo"
    style_sheet(ws, [3, 30, 90, 40])
    title_band(ws, "SCALE PLATES — Client Onboarding", "Your Zomato data in -> insights and action items out. No coding, no waiting.", 1)
    steps = [
        ("STEP 1", "Paste your Payout data", "Open your Zomato settlement report (any month) and go to the 'Order Level' tab. "
         "Select ALL rows (Ctrl+A), copy, then click the 'PayoutDump' sheet and paste into cell A1. "
         "Repeat for each month — paste one report after another (the file keeps the Order Level headers).", ORANGE),
        ("STEP 2", "Paste your Funnel data", "Download your Zomato Business Report (daily, CSV) for the same months. Open it in a text editor, "
         "copy everything, then click the 'FunnelDump' sheet and paste into cell A1. Repeat for each month.", ORANGE),
        ("STEP 3", "Generate insights", "Run this in the folder with the file:   python build_onboarding.py Scale-Plates-Onboarding-Template.xlsx\n"
         "The 'Insights' sheet fills with your health score, key numbers and action items — show it to the client on the spot.", GREEN),
    ]
    r = 4
    for tag, t, d, col in steps:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c = ws.cell(row=r, column=2, value=t)
        c.font = Font(bold=True, size=14, color=col)
        ws.row_dimensions[r].height = 20
        r += 1
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=d)
        c.font = Font(size=11, color="333333")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 78
        r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(row=r, column=2, value="The dump sheets contain SAMPLE data so you can test the flow right now. Replace it with the client's data before the meeting.")
    c.font = Font(bold=True, size=11, color=RED)
    r += 2
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    c = ws.cell(row=r, column=2, value="Need the template in Bengali / Hindi / Malayalam? The Insights sheet is the client-facing page — we can translate it on request.")
    c.font = Font(size=10, color=GREY)

    # ---- PayoutDump ----
    ws = wb.create_sheet("PayoutDump")
    style_sheet(ws, [10] + [22] * 60)
    title_band(ws, "PAYOUT DUMP", "Paste the Zomato 'Order Level' tab here — headers and ALL rows, one report after another. SAMPLE rows below (delete them when you paste).", 1)
    hdr, srows = sample_order_rows()
    for j, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
    for i, r in enumerate(srows, 5):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(size=9, color=GREY)

    # ---- FunnelDump ----
    ws = wb.create_sheet("FunnelDump")
    style_sheet(ws, [16, 22, 14, 12, 18, 26] + [14] * 60)
    title_band(ws, "FUNNEL DUMP", "Paste the daily funnel CSV(s) here — headers and ALL rows, one file after another. SAMPLE rows below (delete them when you paste).", 1)
    fhdr, frows = sample_funnel_rows()
    for j, h in enumerate(fhdr, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = Font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY)
    for i, r in enumerate(frows, 5):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(size=9, color=GREY)

    # ---- Insights (filled by build_onboarding.py) ----
    ws = wb.create_sheet("Insights")
    style_sheet(ws, [24, 16, 60, 40])
    title_band(ws, "INSIGHTS & ACTION ITEMS", "Generated by build_onboarding.py — this sheet fills automatically after Step 3.", 1)
    c = ws.cell(row=4, column=1, value="Run  python build_onboarding.py  on this file to generate insights.")
    c.font = Font(size=12, italic=True, color=GREY)

    wb.save(path)
    print("template written:", path)


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Scale-Plates-Onboarding-Template.xlsx")
    build(out)