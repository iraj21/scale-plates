#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Fill the Scale Plates onboarding template with insights + action items.

Reads PayoutDump and FunnelDump from a template workbook (each may contain
SEVERAL pasted reports — we detect the header rows and split blocks), runs the
Scale Plates pipeline for the latest month, and writes a styled Insights sheet.

Usage:
  python build_onboarding.py Scale-Plates-Onboarding-Template.xlsx [out.xlsx]
"""
import io
import os
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from sp.ingest_payout import norm, find_col, find_fee_col, find_fulfil_col, fval, sval
from sp.ingest_funnel import _parse_date, _num, _mean
from sp.model import kpis, health_score, track, mom
from sp.insights import generate, actionables

NAVY = "1F4E79"
NAVY_D = "143352"
ORANGE = "E67E22"
GREEN = "2E7D32"
RED = "C0392B"
LIGHT = "F4F6F8"
GREY = "6B7280"
WHITE = "FFFFFF"
GOLD = "FFC107"


def rows_of(ws, max_row=None):
    out = []
    for row in ws.iter_rows(values_only=True):
        if max_row and len(out) >= max_row:
            break
        if all(v is None for v in row):
            continue
        out.append(list(row))
    return out


def split_blocks(rows, is_header):
    """Split a dump into blocks at header rows. Returns list of blocks (each a
    list of rows starting with its header row)."""
    blocks, cur = [], None
    for r in rows:
        if is_header(r):
            if cur:
                blocks.append(cur)
            cur = [r]
        elif cur is not None:
            cur.append(r)
    if cur:
        blocks.append(cur)
    return blocks


def is_payout_header(r):
    return any(v and "Subtotal" in str(v) for v in r[:16])


def is_funnel_header(r):
    return len(r) > 5 and str(r[5] or "").strip().lower() == "metric"


# ---------------------------------------------------------------------------
# payout block -> {month: payout dict}
# ---------------------------------------------------------------------------
def parse_payout_block(rows):
    hdr = rows[0]
    cols = {}
    for j, v in enumerate(hdr, 1):
        if v:
            cols[norm(str(v))] = j
    C = dict(
        oid=find_col(cols, "order id"), date=find_col(cols, "order date"),
        status=find_col(cols, "order status"), res=find_col(cols, "res name"),
        resid=find_col(cols, "res id"), subtotal=find_col(cols, "subtotal"),
        packaging=find_col(cols, "packaging"),
        promo=find_col(cols, "restaurant discount promo"),
        bogo=find_col(cols, "restaurant discount bogo"),
        deldisc=find_col(cols, "delivery charge discount"),
        gst=find_col(cols, "total gst collected"), nov=find_col(cols, "net order value"),
        comm=find_fee_col(cols), dist=find_col(cols, "actual order distance"),
        ld=find_col(cols, "long distance enablement fee"), fulfil=find_fulfil_col(cols),
        pm=find_col(cols, "payment mechanism fee"), taxfees=find_col(cols, "taxes on service"),
        tds=find_col(cols, "tds 194o"), payout=find_col(cols, "order level payout"),
    )
    dist_fee_col = C["ld"] or C["fulfil"]
    if C["oid"] is None:
        return {}
    t = dict(orders=0, cancelled=0, subtotal=0.0, packaging=0.0, gst=0.0, nov=0.0,
             commission=0.0, dist_fee=0.0, dist_fee_orders=0, pm=0.0, tax_fees=0.0,
             tds=0.0, payout=0.0, promo_disc=0.0, bogo_disc=0.0, deldisc=0.0,
             dist_km=0.0, promo_orders=0)
    hourly = {}
    daily = {}
    res_name = res_id = None
    for row in rows[1:]:
        oid = sval(row, C["oid"])
        if not oid or oid.startswith("#"):
            continue
        st = sval(row, C["status"]).lower()
        if st in ("cancelled", "rejected"):
            t["cancelled"] += 1
            continue
        if st != "delivered":
            continue
        if res_name is None:
            res_name = sval(row, C["res"])
            res_id = sval(row, C["resid"])
        t["orders"] += 1
        for k, col in [("subtotal", C["subtotal"]), ("packaging", C["packaging"]),
                       ("gst", C["gst"]), ("nov", C["nov"]), ("commission", C["comm"]),
                       ("pm", C["pm"]), ("tax_fees", C["taxfees"]), ("tds", C["tds"]),
                       ("payout", C["payout"]), ("promo_disc", C["promo"]),
                       ("bogo_disc", C["bogo"]), ("deldisc", C["deldisc"])]:
            t[k] += fval(row, col)
        dfv = fval(row, dist_fee_col)
        t["dist_fee"] += dfv
        if dfv > 0:
            t["dist_fee_orders"] += 1
        t["dist_km"] += fval(row, C["dist"])
        if fval(row, C["promo"]) > 0 or fval(row, C["bogo"]) > 0:
            t["promo_orders"] += 1
        dv = row[C["date"] - 1] if C["date"] and len(row) >= C["date"] else None
        try:
            if isinstance(dv, datetime):
                dt = dv
            else:
                s = str(dv or "").strip()
                if len(s) < 19:
                    s = s + " 00:00:00"
                dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        hourly[dt.strftime("%H")] = hourly.get(dt.strftime("%H"), 0) + 1
        daily[dt.strftime("%Y-%m-%d")] = daily.get(dt.strftime("%Y-%m-%d"), 0) + 1
    if t["orders"] == 0:
        return {}
    n = t["orders"]
    fees = t["commission"] + t["dist_fee"] + t["pm"] + t["tax_fees"] + t["tds"]
    dinner = sum(v for k, v in hourly.items() if k in ("19", "20", "21"))
    lunch = sum(v for k, v in hourly.items() if k in ("12", "13", "14"))
    m = min(daily.keys()) if daily else ""
    month = datetime.strptime(m, "%Y-%m-%d").strftime("%Y-%m") if m else ""
    return {month: {
        "restaurant": res_name, "res_id": res_id, "month": month,
        "orders": n, "cancelled": t["cancelled"],
        "subtotal": round(t["subtotal"], 2), "packaging": round(t["packaging"], 2),
        "gst": round(t["gst"], 2), "nov": round(t["nov"], 2),
        "commission": round(t["commission"], 2), "dist_fee": round(t["dist_fee"], 2),
        "dist_fee_orders": t["dist_fee_orders"],
        "payment_mech": round(t["pm"], 2), "tax_on_fees": round(t["tax_fees"], 2),
        "tds": round(t["tds"], 2), "payout": round(t["payout"], 2),
        "promo_disc": round(t["promo_disc"], 2), "bogo_disc": round(t["bogo_disc"], 2),
        "delivery_disc": round(t["deldisc"], 2),
        "avg_distance_km": round(t["dist_km"] / n, 2) if n else 0,
        "aov": round(t["subtotal"] / n, 2) if n else 0,
        "take_rate_pct": round(100 * fees / t["subtotal"], 1) if t["subtotal"] else 0,
        "dist_fee_exposure_pct": round(100 * t["dist_fee_orders"] / n, 1) if n else 0,
        "discount_rate_pct": round(100 * (t["promo_disc"] + t["bogo_disc"]) / t["subtotal"], 1) if t["subtotal"] else 0,
        "payout_pct_of_nov": round(100 * t["payout"] / t["nov"], 1) if t["nov"] else 0,
        "dinner_pct": round(100 * dinner / n, 1) if n else 0,
        "lunch_pct": round(100 * lunch / n, 1) if n else 0,
    }}


# ---------------------------------------------------------------------------
# funnel block -> {month: funnel dict}
# ---------------------------------------------------------------------------
def parse_funnel_block(rows):
    hdr = rows[0]
    date_cols = [_parse_date(str(c)) for c in hdr[5:]]
    col_month = [d.strftime("%Y-%m") if d else None for d in date_cols]
    months = {}
    for row in rows[1:]:
        if len(row) < 6:
            continue
        overview, metric = (row[4] or "").strip(), (row[5] or "").strip()
        res_id, res_name = str(row[0] or ""), str(row[1] or "")
        city, sub = str(row[3] or ""), str(row[2] or "")
        for i, m in enumerate(col_month):
            if m is None:
                continue
            idx = 6 + i
            if idx >= len(row):
                continue
            val = _num(row[idx])
            mo = months.setdefault(m, {
                "restaurant": res_name, "res_id": res_id, "city": city, "subzone": sub,
                "month": m, "days": 0, "zero_order_days": 0,
                "orders": 0.0, "sales": 0.0, "impressions": 0.0, "menu_opens": 0.0,
                "cart_builds": 0.0, "placed_orders": 0.0,
                "ad_spend": 0.0, "sales_from_ads": 0.0, "ads_orders": 0.0,
                "ads_ctr": [], "rating": [], "for_acc": [], "kpt": [], "online_pct": [],
                "repeat_orders": 0.0, "new_orders": 0.0, "lapsed_orders": 0.0,
                "breakfast": 0.0, "lunch": 0.0, "snacks": 0.0, "dinner": 0.0, "late_night": 0.0,
                "complaints": 0.0, "lost_sales": 0.0, "bad_orders": 0.0, "rejected": 0.0,
                "day_orders": {},
            })
            key = (overview, metric)
            if key == ("Sales", "Delivered orders"):
                mo["orders"] += val; mo["day_orders"][i] = val
            elif key == ("Sales", "Sales (Rs)"):
                mo["sales"] += val
            elif key == ("Customer funnel", "Impressions"):
                mo["impressions"] += val
            elif key == ("Customer funnel", "Menu opens"):
                mo["menu_opens"] += val
            elif key == ("Customer funnel", "Cart builds"):
                mo["cart_builds"] += val
            elif key == ("Customer funnel", "Placed Orders"):
                mo["placed_orders"] += val
            elif key == ("Ads", "Ads spend (Rs)"):
                mo["ad_spend"] += val
            elif key == ("Ads", "Sales from ads (Rs)"):
                mo["sales_from_ads"] += val
            elif key == ("Ads", "Ads orders"):
                mo["ads_orders"] += val
            elif key == ("Ads", "Ads CTR (%)"):
                mo["ads_ctr"].append(val)
            elif key == ("Customer experience", "Average rating"):
                mo["rating"].append(val)
            elif key == ("Customer experience", "FOR accuracy (%)"):
                mo["for_acc"].append(val)
            elif key == ("Customer experience", "KPT (in minutes)"):
                mo["kpt"].append(val)
            elif key == ("Customer experience", "Online %"):
                mo["online_pct"].append(val)
            elif key == ("Customer segmentation", "Repeat user orders"):
                mo["repeat_orders"] += val
            elif key == ("Customer segmentation", "New user orders"):
                mo["new_orders"] += val
            elif key == ("Customer segmentation", "Lapsed user orders"):
                mo["lapsed_orders"] += val
            elif key == ("Customer segmentation", "Breakfast orders"):
                mo["breakfast"] += val
            elif key == ("Customer segmentation", "Lunch orders"):
                mo["lunch"] += val
            elif key == ("Customer segmentation", "Snacks orders"):
                mo["snacks"] += val
            elif key == ("Customer segmentation", "Dinner orders"):
                mo["dinner"] += val
            elif key == ("Customer segmentation", "Late night orders"):
                mo["late_night"] += val
            elif key == ("Customer experience", "Total complaints"):
                mo["complaints"] += val
            elif key == ("Customer experience", "Lost sales (Rs)"):
                mo["lost_sales"] += val
            elif key == ("Customer experience", "Bad orders"):
                mo["bad_orders"] += val
            elif key == ("Customer experience", "Rejected orders"):
                mo["rejected"] += val
    out = {}
    for m, mo in months.items():
        n = mo["orders"]
        mo["days"] = len(mo["day_orders"])
        mo["zero_order_days"] = sum(1 for v in mo["day_orders"].values() if v == 0)
        mo["aov"] = round(mo["sales"] / n, 2) if n else 0
        mo["roas"] = round(mo["sales_from_ads"] / mo["ad_spend"], 2) if mo["ad_spend"] else 0
        mo["ad_dependency_pct"] = round(100 * mo["ads_orders"] / n, 1) if n else 0
        mo["ad_cost_per_order"] = round(mo["ad_spend"] / mo["ads_orders"], 2) if mo["ads_orders"] else 0
        mo["ctr_pct"] = round(_mean(mo["ads_ctr"]), 1)
        mo["rating"] = round(_mean(mo["rating"]), 2)
        mo["for_accuracy_pct"] = round(_mean(mo["for_acc"]), 1)
        mo["kpt_min"] = round(_mean(mo["kpt"]), 1)
        mo["online_pct"] = round(_mean(mo["online_pct"]), 1)
        mo["repeat_rate_pct"] = round(100 * mo["repeat_orders"] / n, 1) if n else 0
        mo["new_pct"] = round(100 * mo["new_orders"] / n, 1) if n else 0
        mo["lapsed_pct"] = round(100 * mo["lapsed_orders"] / n, 1) if n else 0
        mo["i2m_pct"] = round(100 * mo["menu_opens"] / mo["impressions"], 1) if mo["impressions"] else 0
        mo["m2c_pct"] = round(100 * mo["cart_builds"] / mo["menu_opens"], 1) if mo["menu_opens"] else 0
        mo["c2o_pct"] = round(100 * mo["placed_orders"] / mo["cart_builds"], 1) if mo["cart_builds"] else 0
        mo["dinner_pct"] = round(100 * mo["dinner"] / n, 1) if n else 0
        mo["lunch_pct"] = round(100 * mo["lunch"] / n, 1) if n else 0
        mo["breakfast_pct"] = round(100 * mo["breakfast"] / n, 1) if n else 0
        mo["snacks_pct"] = round(100 * mo["snacks"] / n, 1) if n else 0
        mo["late_night_pct"] = round(100 * mo["late_night"] / n, 1) if n else 0
        mo["complaints_per_100"] = round(100 * mo["complaints"] / n, 1) if n else 0
        mo["bad_order_pct"] = round(100 * mo["bad_orders"] / n, 1) if n else 0
        mo["rejected_pct"] = round(100 * mo["rejected"] / n, 1) if n else 0
        for k in ("ads_ctr", "for_acc", "kpt", "online_pct", "day_orders"):
            mo.pop(k, None)
        out[m] = mo
    return out


# ---------------------------------------------------------------------------
# workbook fill
# ---------------------------------------------------------------------------
def fill_insights(wb, report):
    ws = wb["Insights"]
    idx = wb.sheetnames.index("Insights")
    del wb["Insights"]
    ws = wb.create_sheet("Insights", idx)
    for i, w in enumerate([24, 16, 60, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D5DDE5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    k = report["kpis"]
    h = report["health"]

    def put(r, c, v, font=None, fill=None, align=None, border=None):
        cell = ws.cell(row=r, column=c, value=v)
        if font:
            cell.font = font
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if align:
            cell.alignment = align
        if border:
            cell.border = border
        return cell

    def title_band(r, t, sub):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=t)
        c.font = Font(bold=True, size=16, color=WHITE)
        c.fill = PatternFill("solid", fgColor=NAVY_D)
        ws.row_dimensions[r].height = 26
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=4)
        c = ws.cell(row=r + 1, column=1, value=sub)
        c.font = Font(size=10, color=GREY)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        ws.row_dimensions[r + 1].height = 20

    r = 1
    title_band(r, "SCALE PLATES — INSIGHTS & ACTION ITEMS",
               f"{report['restaurant']}  •  {report.get('city','')}  •  {report['month']}")
    r += 3
    put(r, 1, "HEALTH SCORE", Font(bold=True, size=13, color=NAVY))
    put(r, 2, f"{h['overall']} / 100", Font(bold=True, size=13, color=GREEN if h['overall'] >= 75 else ORANGE))
    put(r, 3, report["track"], Font(size=10, bold=True, color=NAVY))
    r += 1
    for name, val in h["subs"].items():
        put(r, 1, name, Font(size=10, color=GREY))
        col = GREEN if val >= 80 else (ORANGE if val >= 60 else RED)
        put(r, 2, val, Font(bold=True, size=10, color=col))
        r += 1
    r += 1

    put(r, 1, "KEY NUMBERS", Font(bold=True, size=13, color=NAVY))
    r += 1
    rows = [
        ("Orders", f"{int(k['orders']):,}", None),
        ("Average order size", f"₹{k['aov']:,.0f}", None),
        ("Sales (menu value)", f"₹{k['subtotal']:,.0f}", None),
        ("Money you received", f"₹{k['order_payout']:,.0f}", None),
        ("Platform take rate", f"{k['take_rate']*100:.1f}%", None),
        ("Ad spend", f"₹{k['ad_spend']:,.0f}", None),
        ("Ad return (ROAS)", f"{k['roas']:.1f}x" if k["roas"] else "—", None),
        ("Orders from ads", f"{k['ad_dependency_pct']:.0f}%", None),
        ("Repeat customers", f"{k['repeat_rate_pct']:.0f}%" if k["repeat_rate_pct"] else "—", None),
        ("Rating", f"{k['rating']:.2f}" if k["rating"] else "—", None),
        ("Hidden distance-fee orders", f"{k['ld_exposure']*100:.0f}%", None),
        ("Zero-order days", f"{int(k['zero_order_days'])}", None),
    ]
    for name, val, _ in rows:
        put(r, 1, name, Font(size=10, color=GREY), border=border)
        put(r, 2, val, Font(bold=True, size=10, color=NAVY), border=border)
        r += 1
    r += 1

    put(r, 1, "INSIGHTS & ACTION ITEMS", Font(bold=True, size=13, color=NAVY))
    r += 1
    for it in report["insights"]:
        pcol = RED if it["priority"] == "P0" else (ORANGE if it["priority"] == "P1" else GREY)
        put(r, 1, it["priority"], Font(bold=True, size=10, color=WHITE), fill=pcol)
        put(r, 2, it["category"], Font(size=9, color=GREY))
        put(r, 3, it["title"], Font(bold=True, size=10, color=NAVY))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 18
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = put(r, 1, it["detail"] + "  →  " + it["recommendation"],
                Font(size=10, color="333333"))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 34
        r += 2
    put(r, 1, "Generated by Scale Plates on " + datetime.now().strftime("%d %b %Y %H:%M"),
        Font(size=9, italic=True, color=GREY))


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    dst = sys.argv[2] if len(sys.argv) > 2 else src.replace(".xlsx", "-insights.xlsx")
    wb = load_workbook(src)
    payout_rows = rows_of(wb["PayoutDump"])
    funnel_rows = rows_of(wb["FunnelDump"])

    payouts = {}
    for block in split_blocks(payout_rows, is_payout_header):
        payouts.update(parse_payout_block(block))
    funnels = {}
    for block in split_blocks(funnel_rows, is_funnel_header):
        funnels.update(parse_funnel_block(block))

    if not payouts:
        print("ERROR: no payout data found in PayoutDump. Paste the Zomato 'Order Level' tab (headers + rows).")
        sys.exit(2)
    pmonths = sorted(payouts)
    latest = pmonths[-1]
    prev = pmonths[-2] if len(pmonths) > 1 else None
    if latest not in funnels:
        print("WARNING: no funnel data for month", latest, "- report will be payout-only (Track 1).")
    entry = {"payout": payouts.get(latest), "funnel": funnels.get(latest), "month": latest,
             "restaurant": (payouts.get(latest) or {}).get("restaurant", "Restaurant"),
             "res_id": (payouts.get(latest) or {}).get("res_id", ""),
             "city": (funnels.get(latest) or {}).get("city", "")}
    prev_entry = {"payout": payouts.get(prev), "funnel": funnels.get(prev), "month": prev} if prev else None

    k = kpis(entry["payout"], entry["funnel"])
    ins = generate(entry, prev_entry)
    report = {
        "restaurant": entry["restaurant"], "res_id": entry["res_id"], "city": entry["city"],
        "platform": "ZOMATO", "month": latest,
        "kpis": k, "mom": mom(entry, prev_entry) if prev_entry else None,
        "health": health_score(k), "track": track(k),
        "insights": ins, "actionables": actionables(ins),
    }
    fill_insights(wb, report)
    wb.save(dst)
    print("filled insights:", dst)
    print("restaurant:", report["restaurant"], "| month:", latest, "| health:", report["health"]["overall"])
    print("months parsed:", pmonths)


if __name__ == "__main__":
    main()